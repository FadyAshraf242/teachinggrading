import sqlite3
from flask import session, Flask, render_template, request, redirect, url_for, flash, send_file, jsonify
from flask_bcrypt import Bcrypt
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
import os
import logging
import pytesseract  
from PIL import Image
from werkzeug.utils import secure_filename
from datetime import timedelta, datetime
import numpy as np
import requests
import json
import pandas as pd
from flask import send_file
import pandas as pd
from flask import send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment
from docx import Document
app = Flask(__name__)
app.secret_key = "Delta_U_X"
bcrypt = Bcrypt(app)
login_manager = LoginManager()
login_manager.login_view = "login"
login_manager.init_app(app)
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(minutes=10)  
logging.basicConfig(level=logging.DEBUG)
SECRET_CODE = "hossam"
app.config['MAX_CONTENT_LENGTH'] = 10 * 1024 * 1024
UPLOAD_FOLDER = 'uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS
@app.before_request
def clear_sessions_on_restart():
    if not app.config.get('SESSIONS_CLEARED', False):
        session.clear()
        app.config['SESSIONS_CLEARED'] = True
@app.before_request
def check_session_timeout():
    if 'last_activity' in session:
        try:
            now = datetime.now()
            last_activity = datetime.strptime(session['last_activity'], "%Y-%m-%d %H:%M:%S")
            session_timeout = app.config.get('PERMANENT_SESSION_LIFETIME', timedelta(minutes=10))
            if (now - last_activity) > session_timeout:
                session.clear()
                flash("Session expired due to inactivity. Please log in again.", "danger")
                return redirect(url_for("login"))
        except (KeyError, ValueError):
            session.clear()
            flash("Session expired or invalid. Please log in again.", "danger")
            return redirect(url_for("login"))
import sqlite3
def init_db():
    with sqlite3.connect("database.db") as conn:
        cursor = conn.cursor()
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS Std (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL UNIQUE,
            email TEXT NOT NULL UNIQUE,
            password TEXT NOT NULL,
            role TEXT NOT NULL CHECK(role IN ('student', 'professor'))
        );''')
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS Essays (
            essay_id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            title TEXT,
            max_score REAL,
            ideal_answer TEXT,
            FOREIGN KEY(user_id) REFERENCES Std(id)
        );''')
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS EssaySubmissions (
            submission_id INTEGER PRIMARY KEY AUTOINCREMENT,
            essay_id INTEGER NOT NULL,
            user_id INTEGER NOT NULL,
            written_answer TEXT,
            submission_date DATETIME DEFAULT CURRENT_TIMESTAMP,
            score REAL,
            evaluation TEXT,
            status TEXT DEFAULT 'Not Submitted',
            FOREIGN KEY(essay_id) REFERENCES Essays(essay_id),
            FOREIGN KEY(user_id) REFERENCES Std(id)
        );''')
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS EssayQuestions (
            question_id INTEGER PRIMARY KEY AUTOINCREMENT,
            essay_id INTEGER NOT NULL,
            question TEXT NOT NULL,
            ideal_answer TEXT NOT NULL,
            score REAL DEFAULT 0,
            FOREIGN KEY(essay_id) REFERENCES Essays(essay_id)
        );''')
        cursor.execute("PRAGMA table_info(EssaySubmissions);")
        columns = [col[1] for col in cursor.fetchall()]
        if "grading_status" not in columns:
            cursor.execute("ALTER TABLE EssaySubmissions ADD COLUMN grading_status TEXT DEFAULT 'Not Graded';")
            print("Column 'grading_status' added successfully.")
        cursor.execute("PRAGMA table_info(Essays);")
        columns = [col[1] for col in cursor.fetchall()]
        if "review_enabled" not in columns:
            cursor.execute("ALTER TABLE Essays ADD COLUMN review_enabled INTEGER DEFAULT 0;")
            print("Column 'review_enabled' added successfully.")
        if "review_time" not in columns:
            cursor.execute("ALTER TABLE Essays ADD COLUMN review_time DATETIME DEFAULT NULL;")
            print("Column 'review_time' added successfully.")
    conn.commit()
    print("Database and tables updated successfully!")
class User(UserMixin):
    def __init__(self, id, username, email, role):
        self.id = id
        self.username = username
        self.email = email
        self.role = role
@login_manager.user_loader
def load_user(user_id):
    with sqlite3.connect("database.db") as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM Std WHERE id = ?", (user_id,))
        user = cursor.fetchone()
        if user:
            return User(id=user[0], username=user[1], email=user[2], role=user[4])
        return None
@app.route("/")
def home():
    return render_template("Home.html")
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email = request.form["email"]
        password = request.form["password"]
        with sqlite3.connect("database.db") as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM Std WHERE email = ?", (email,))
            user = cursor.fetchone()
            if user and bcrypt.check_password_hash(user[3], password):
                role = user[4]
                user_obj = User(id=user[0], username=user[1], email=user[2], role=role)
                login_user(user_obj)
                session["user_id"] = user_obj.id
                flash("Login successful", "success")
                if role == "professor":
                    return redirect(url_for("professor_dashboard"))
                elif role == "student":
                    return redirect(url_for("student_dashboard"))
            else:
                flash("Invalid email or password", "danger")
    return render_template("auth/Login.html")
@app.route("/signup", methods=["GET", "POST"])
def signup():
    if request.method == "POST":
        username = request.form["username"]
        email = request.form["email"]
        password = request.form["password"]
        confirm_password = request.form["confirm_password"]
        role = request.form["user_type"]
        secret_code = request.form.get("secret_code", "").strip()
        if password != confirm_password:
            flash("Passwords do not match.", "danger")
            return redirect(request.url)
        if role == "professor" and secret_code != SECRET_CODE:
            flash("Invalid secret code for Professor signup.", "danger")
            return redirect(request.url)
        password_hash = bcrypt.generate_password_hash(password).decode("utf-8")
        with sqlite3.connect("database.db") as conn:
            cursor = conn.cursor()
            cursor.execute("INSERT INTO Std (username, email, password, role) VALUES (?, ?, ?, ?)",
                           (username, email, password_hash, role))
            conn.commit()
        flash("Account created successfully. Please log in.", "success")
        return redirect(url_for("login"))
    return render_template("auth/Signup.html")
@app.route("/update_score", methods=["POST"])
@login_required
def update_score():
    if current_user.role != 'professor':
        flash("Access denied.", "danger")
        return redirect(url_for("home"))
    submission_id = request.form.get("submission_id")
    new_score = request.form.get("score")
    new_evaluation = request.form.get("evaluation")
    if not submission_id or not new_score or not new_evaluation:
        flash("Missing data. Please provide all required fields.", "danger")
        return redirect(url_for("professor_dashboard"))
    try:
        with sqlite3.connect("database.db") as conn:
            cursor = conn.cursor()
            cursor.execute("""
                UPDATE EssaySubmissions
                SET score = ?, evaluation = ?, grading_status = 'Graded'
                WHERE submission_id = ?
            """, (new_score, new_evaluation, submission_id))
            conn.commit()
        flash("Score updated successfully!", "success")
    except Exception as e:
        print(f"Error updating score: {e}")
        flash("An error occurred while updating the score.", "danger")
    return redirect(url_for("professor_dashboard"))
@app.route("/professor_dashboard")
@login_required
def professor_dashboard():
    if current_user.role != 'professor':
        flash("Access denied.", "danger")
        return redirect(url_for("home"))
    with sqlite3.connect("database.db") as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT essay_id, title FROM Essays WHERE user_id = ?", (current_user.id,))
        quizzes = cursor.fetchall()
        submissions = {}
        for quiz in quizzes:
            quiz_id = quiz[0]
            cursor.execute("""
                SELECT S.username, ES.score, ES.evaluation, ES.submission_id, ES.grading_status
                FROM EssaySubmissions ES
                JOIN Std S ON ES.user_id = S.id
                WHERE ES.essay_id = ?
            """, (quiz_id,))
            submissions[quiz_id] = cursor.fetchall()
    return render_template("dashboard/professor_dashboard.html", quizzes=quizzes, submissions=submissions)
@app.route("/download_excel/<int:quiz_id>")
@login_required
def download_excel(quiz_id):
    with sqlite3.connect("database.db") as conn:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT S.username, ES.score, ES.evaluation
            FROM EssaySubmissions ES
            JOIN Std S ON ES.user_id = S.id
            WHERE ES.essay_id = ?
        """, (quiz_id,))
        data = cursor.fetchall()
    df = pd.DataFrame(data, columns=["Student Name", "Score", "Evaluation"])
    file_path = f"quiz_{quiz_id}_results.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f"Quiz {quiz_id} Results"
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")  
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=col_name)
        cell.alignment = Alignment(horizontal="center", vertical="center")  
        sheet.column_dimensions[cell.column_letter].width = max(len(col_name), 15)  
    for col_idx, col_name in enumerate(df.columns, start=1):
        max_length = max((len(str(sheet.cell(row=row, column=col_idx).value)) for row in range(1, sheet.max_row + 1)), default=0)
        sheet.column_dimensions[sheet.cell(row=1, column=col_idx).column_letter].width = max_length + 2
    workbook.save(file_path)
    return send_file(file_path, as_attachment=True)
@app.route("/edit_quiz", methods=["GET", "POST"])
@login_required
def edit_quiz():
    if request.method == "POST":
        essay_id = request.form["essay_id"]
        title = request.form["title"]
        max_score = float(request.form["max_score"])
        max_score_display = request.form["max_score_display"]
        instructions = request.form["instructions"]
        ideal_answer = request.form["ideal_answer"]
        cursor.execute("""
         SELECT E.essay_id, E.title, E.max_score, E.instructions FROM Essays E WHERE E.essay_id = ? """, (essay_id,))
        flash("Quiz/Essay details updated successfully!", "success")
        return redirect(url_for("professor_dashboard"))
    essay_id = request.args.get("essay_id")
    with sqlite3.connect("database.db") as conn:
        cursor = conn.cursor()
        cursor.execute("""SELECT E.essay_id, E.title, E.max_score, E.instructions
            FROM Essays E
            WHERE E.essay_id = ?""", (essay_id,))
        quiz_data = cursor.fetchone()
    if quiz_data is None:
        flash("Quiz/Essay not found.", "danger")
        return redirect(url_for("professor_dashboard"))
    max_score_display = f"{quiz_data[2]}/10"
    return render_template("dashboard/edit_quiz.html", quiz_data=quiz_data, max_score_display=max_score_display)
GEMINI_API_KEY = "AIzaSyAjiN6AKNgdTqyb26Wvb4RPO31mX5pO-7M"
def extract_text_from_docx(file_path):
    doc = Document(file_path)
    text = "\n".join([p.text for p in doc.paragraphs if p.text.strip()])
    return text
def parse_quiz_from_gemini(text):
    prompt = f"""
        Extract questions, model answers, and scores from the given text.
        Format output strictly as: [[question, answer, score], [question, answer, score], ...].
        If no score is mentioned, assign 10 by default.
        Text: {text}
    """
    url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash:generateContent?key={GEMINI_API_KEY}"
    headers = {"Content-Type": "application/json"}
    payload = {
        "contents": [{"parts": [{"text": prompt}]}]
    }
    response = requests.post(url, headers=headers, json=payload)
    if response.status_code == 200:
        try:
            raw_output = response.json()["candidates"][0]["content"]["parts"][0]["text"]
            parsed_output = json.loads(raw_output.replace("```", "").replace("json", "").strip())
            return parsed_output
        except Exception as e:
            print(f"Error parsing AI response: {e}")
            return []
    return []
@app.route("/extract_questions", methods=["POST"])
@login_required
def extract_questions():
    if "quiz_file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    file = request.files["quiz_file"]
    if file.filename == "":
        return jsonify({"error": "No selected file"}), 400
    if file and file.filename.endswith(".docx"):
        file_path = os.path.join(app.config["UPLOAD_FOLDER"], file.filename)
        file.save(file_path)
        extracted_text = extract_text_from_docx(file_path)
        questions_data = parse_quiz_from_gemini(extracted_text)
        os.remove(file_path)  
        return jsonify(questions_data)
    return jsonify({"error": "Invalid file format"}), 400
@app.route("/create_quiz", methods=["GET", "POST"])
@login_required
def create_quiz():
    if current_user.role != 'professor':
        flash("Access denied.", "danger")
        return redirect(url_for("home"))
    if request.method == "POST":
        title = request.form["title"]
        max_score = float(request.form["max_score"])
        questions = request.form.getlist("questions[]")
        ideal_answers = request.form.getlist("ideal_answers[]")
        question_scores = [float(score) for score in request.form.getlist("question_scores[]")]
        with sqlite3.connect("database.db") as conn:
            cursor = conn.cursor()
            cursor.execute("INSERT INTO Essays (title, max_score, user_id) VALUES (?, ?, ?)", (title, max_score, current_user.id))
            essay_id = cursor.lastrowid
            for i in range(len(questions)):
                cursor.execute("""
                    INSERT INTO EssayQuestions (essay_id, question, ideal_answer, score)
                    VALUES (?, ?, ?, ?)
                """, (essay_id, questions[i], ideal_answers[i], question_scores[i]))
            cursor.execute("SELECT id FROM Std WHERE role = 'student'")
            students = cursor.fetchall()
            for student in students:
                cursor.execute("""
                    INSERT INTO EssaySubmissions (essay_id, user_id, status, grading_status)
                    VALUES (?, ?, 'Not Submitted', 'Not Graded')
                """, (essay_id, student[0]))
            conn.commit()
        flash("Quiz created successfully!", "success")
        return redirect(url_for("professor_dashboard"))
    return render_template("dashboard/create_quiz.html")
@app.route("/student_dashboard")
@login_required
def student_dashboard():
    if current_user.role != 'student':
        flash("You must be a student to access this page.", "danger")
        return redirect(url_for("home"))
    try:
        with sqlite3.connect("database.db") as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT e.essay_id, e.title, es.status, es.grading_status
                FROM Essays e
                LEFT JOIN EssaySubmissions es ON e.essay_id = es.essay_id AND es.user_id = ?
            """, (current_user.id,))
            available_essays = cursor.fetchall()
    except sqlite3.Error as e:
        print("Database error:", e)
        available_essays = []
    return render_template("dashboard/student_dashboard.html", essays=available_essays)
import threading
def submit_essay_background(essay_id, user_id, answers, image_files):
    with sqlite3.connect("database.db") as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT question_id, ideal_answer, score, question FROM EssayQuestions WHERE essay_id = ?", (essay_id,))
        questions_data = cursor.fetchall()
        total_score = 0
        evaluations = []
        processed_answers = []
        for i, (question_id, ideal_answer, question_score, question) in enumerate(questions_data):
            student_answer = answers[i] if i < len(answers) else ""
            image_answer = image_files[i] if i < len(image_files) and image_files[i].filename != "" else None
            if image_answer:
                filename = secure_filename(image_answer.filename)
                file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                image_answer.save(file_path)
                student_answer = process_image_with_google_cloud(file_path).strip()
                os.remove(file_path)
            evaluation_score = generate_evaluation(student_answer, ideal_answer, question_score, question, essay_id, user_id)
            total_score += evaluation_score
            evaluations.append(f"Q{i+1}: {evaluation_score}/{question_score}")
            processed_answers.append(student_answer)
        final_evaluation = " | ".join(evaluations)
        cursor.execute("""
            UPDATE EssaySubmissions
            SET written_answer = ?, score = ?, evaluation = ?, grading_status = 'Graded'
            WHERE essay_id = ? AND user_id = ?
        """, (json.dumps(processed_answers), total_score, final_evaluation, essay_id, user_id))
        conn.commit()
@app.route("/submit_essay/<int:essay_id>", methods=["GET", "POST"])
@login_required
def submit_essay(essay_id):
    if request.method == "POST":
        try:
            answers = request.form.getlist("answers[]")
            image_files = request.files.getlist("image_answers[]")
            with sqlite3.connect("database.db") as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    UPDATE EssaySubmissions
                    SET grading_status = 'Grading', status = 'Submitted'
                    WHERE essay_id = ? AND user_id = ?
                """, (essay_id, current_user.id))
                conn.commit()
            threading.Thread(target=submit_essay_background, args=(essay_id, current_user.id, answers, image_files)).start()
            flash("Your essay has been submitted successfully! Grading in progress...", "success")
            return redirect(url_for("student_dashboard"))
        except Exception as e:
            print(f"Error during essay submission: {e}")
            flash("An error occurred while submitting the essay. Please try again.", "error")
            return redirect(request.url)
    with sqlite3.connect("database.db") as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT title FROM Essays WHERE essay_id = ?", (essay_id,))
        essay = cursor.fetchone()
        cursor.execute("SELECT question_id, question FROM EssayQuestions WHERE essay_id = ?", (essay_id,))
        questions = [{"id": row[0], "question_text": row[1]} for row in cursor.fetchall()]
    if not essay or not questions:
        flash("Essay not found.", "danger")
        return redirect(url_for("student_dashboard"))
    return render_template("dashboard/Submit.html", essay=essay, questions=questions)
def process_written_text(text):
    return text.strip()
import os
from google.cloud import vision
import io
google_credentials_path = "college2-430912-c687b27d3952.json"
if os.path.exists(google_credentials_path):
    client = vision.ImageAnnotatorClient.from_service_account_json(google_credentials_path)
else:
    raise FileNotFoundError("Google Cloud Vision API credentials file not found!")
def process_image_with_google_cloud(file_path):
    """ استخراج النص من الصورة باستخدام Google Cloud Vision API """
    with io.open(file_path, 'rb') as image_file:
        content = image_file.read()
    image = vision.Image(content=content)
    response = client.text_detection(image=image)
    texts = response.text_annotations
    return texts[0].description.strip() if texts else ""
@app.route("/upload_ocr", methods=["POST"])
@login_required
def upload_ocr():
    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400
    file = request.files["file"]
    if file.filename == "":
        return jsonify({"error": "No selected file"}), 400
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)
        extracted_text = process_image_with_google_cloud(file_path)
        os.remove(file_path)
        return jsonify({"extracted_text": extracted_text})
    return jsonify({"error": "Invalid file format"}), 400
import threading
import time
import random
GEMINI_API_KEYS = [
    "AIzaSyAjiN6AKNgdTqyb26Wvb4RPO31mX5pO-7M"
]
request_queue = []
queue_lock = threading.Lock()
requests_count = {key: 0 for key in GEMINI_API_KEYS}
def reset_request_count():
    global requests_count
    while True:
        time.sleep(60)  
        requests_count = {key: 0 for key in GEMINI_API_KEYS}
threading.Thread(target=reset_request_count, daemon=True).start()
def send_to_gemini_api(prompt_text, max_retries=5):
    retries = 0
    while retries < max_retries:
        api_key = random.choice(GEMINI_API_KEYS)
        if requests_count[api_key] < 15:
            url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash:generateContent?key={api_key}"
            headers = {"Content-Type": "application/json"}
            payload = {
                "contents": [
                    {
                        "parts": [
                            {"text": prompt_text}
                        ]
                    }
                ]
            }
            try:
                response = requests.post(url, headers=headers, data=json.dumps(payload))
                if response.status_code == 200:
                    requests_count[api_key] += 1
                    response_data = response.json()
                    evaluation = response_data['candidates'][0]['content']['parts'][0]['text'].strip()
                    return evaluation
                else:
                    print(f"Error: API request failed with status code {response.status_code}")
            except Exception as e:
                print(f"Error: {e}")
        retries += 1
        time.sleep(60)  
    return str("failed")  
def process_queue():
    while True:
        with queue_lock:
            if request_queue:
                prompt_text, callback = request_queue.pop(0)
                evaluation = send_to_gemini_api(prompt_text)
                callback(evaluation)
        time.sleep(1)  
threading.Thread(target=process_queue, daemon=True).start()
def generate_evaluation(written_answer, ideal_answer, question_score, question, essay_id, user_id):
    prompt_text = f"""
        This program is designed to grade a student's answer based on meaning rather than exact wording. The professor provides a question along with an expected correct answer, and the student submits a written response.
        Your task is to **assign a fair score out of {question_score} based on the correctness of the content**, not just textual similarity. If the student's response is **logically valid and answers the question correctly based on a reasonable interpretation**, they should receive full credit ({question_score} points).
        🔹 **If the question asks for factual or personal information (e.g., age, name, hometown), then the student's personal correct response should be fully accepted, even if it differs from the expected answer.**  
        🔹 **If the response is incorrect, off-topic, or does not logically answer the question, deduct points accordingly.**  
        ❗ Do not penalize correct answers just because they are different from the expected answer.  
        Please provide **only the score (out of {question_score})**, without any additional text.
        **Maximum Score:** {question_score}  
        **Question:** {question}  
        **Expected correct answer:** {ideal_answer}  
        **Student's answer:** {written_answer}  
        Return only the numeric score (out of {question_score}).
    """
    def callback(evaluation):
        try:
            score = float(evaluation)
            print(f"this is grade (converted to float): {score}")
            return score
        except ValueError:
            print("Error: AI response was not a valid number.")
            return 0
    evaluation = send_to_gemini_api(prompt_text)
    return callback(evaluation)
def save_submission(cursor, essay_id, answer_text):
    cursor.execute("""
        SELECT submission_id 
        FROM EssaySubmissions 
        WHERE essay_id = ? AND user_id = ? AND status = 'Not Submitted'
    """, (essay_id, current_user.id))
    previous_submission = cursor.fetchone()
    if previous_submission:
        submission_id = previous_submission[0]
        cursor.execute("""
            UPDATE EssaySubmissions
            SET written_answer = ?, submission_date = ? ,status = 'Submitted'
            WHERE submission_id = ?
        """, (answer_text, datetime.now(), submission_id))
    else:
        cursor.execute("""
            INSERT INTO EssaySubmissions (essay_id, user_id, written_answer, submission_date, status)
            VALUES (?, ?, ?, ?, 'Submitted')
        """, (essay_id, current_user.id, answer_text, datetime.now()))
def fetch_essay_content(essay_id):
    with sqlite3.connect("database.db") as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM Essays WHERE essay_id = ?", (essay_id,))
        return cursor.fetchone()
def save_uploaded_file(file):
    filename = secure_filename(file.filename)
    file_path = os.path.join(UPLOAD_FOLDER, filename)
    file.save(file_path)
    return file_path
@app.route("/check_essay_quiz/<int:essay_id>")
@login_required
def check_essay_quiz(essay_id):
    with sqlite3.connect("database.db") as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM Essays WHERE essay_id = ?", (essay_id,))
        essay = cursor.fetchone()
    if not essay:
        flash("Essay not found", "danger")
        return redirect(url_for("student_dashboard"))
    return render_template("check_essay_quiz.html", essay=essay)
@app.route("/toggle_review", methods=["POST"])
@login_required
def toggle_review():
    if current_user.role != 'professor':
        return jsonify({"error": "Access denied"}), 403
    essay_id = request.form.get("essay_id")
    review_enabled = request.form.get("review_enabled")
    review_time = request.form.get("review_time")
    with sqlite3.connect("database.db") as conn:
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE Essays SET review_enabled = ?, review_time = ?
            WHERE essay_id = ?
        """, (review_enabled, review_time, essay_id))
        conn.commit()
    return jsonify({"message": "Review settings updated successfully!"})
@app.route("/review_essay/<int:essay_id>")
@login_required
def review_essay(essay_id):
    with sqlite3.connect("database.db") as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT review_enabled, review_time, max_score FROM Essays WHERE essay_id = ?", (essay_id,))
        review_settings = cursor.fetchone()
        if review_settings:
            review_enabled = review_settings[0]
            review_time = review_settings[1]
            max_score = review_settings[2]  
            if review_time:
                review_time = datetime.strptime(review_time, "%Y-%m-%dT%H:%M")
            if not review_enabled and (not review_time or datetime.now() < review_time):
                flash("Review is not available yet.", "danger")
                return redirect(url_for("student_dashboard"))
        cursor.execute("""
            SELECT es.written_answer, es.submission_date, es.status, es.score, es.evaluation, es.grading_status
            FROM EssaySubmissions es
            WHERE es.essay_id = ? AND es.user_id = ?
        """, (essay_id, current_user.id))
        submitted_essay = cursor.fetchone()
        if not submitted_essay:
            flash("You have not submitted this essay yet.", "danger")
            return redirect(url_for("student_dashboard"))
        cursor.execute("""
            SELECT question
            FROM EssayQuestions
            WHERE essay_id = ?
        """, (essay_id,))
        questions = [row[0] for row in cursor.fetchall()]
        student_answers = json.loads(submitted_essay[0]) if submitted_essay[0] else []
        questions_with_answers = [
            {"question": questions[i], "answer": student_answers[i]} if i < len(student_answers) else {"question": questions[i], "answer": "No answer provided"}
            for i in range(len(questions))
        ]
        essay_details = {
            "submission_date": submitted_essay[1],
            "status": submitted_essay[2],
            "score": submitted_essay[3],
            "evaluation": submitted_essay[4],
            "grading_status": submitted_essay[5],
            "max_score": max_score,  
            "questions_with_answers": questions_with_answers
        }
    return render_template("dashboard/review_essay.html", essay_details=essay_details)
@app.route("/logout")
@login_required
def logout():
    logout_user()
    flash("You have been logged out.", "info")
    return redirect(url_for("home"))    
if __name__ == "__main__":
    init_db()
    app.run(host='0.0.0.0', port=8080)
